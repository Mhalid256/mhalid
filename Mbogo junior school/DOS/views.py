from django.http import HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import render
from . models import student
from . models import reportcard
from.forms import studentform
from django.http import HttpResponse
from django.http import FileResponse
import io
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
import openpyxl

# Create your views here.
def index(request):
    return render(request,'DOS/index.html',{
        'DOS':student.objects.all()
    })
def view_student(request,id):
        student=student.objects.get(pk=id)
        return HttpResponseRedirect(reverse('index'))
    
def add(request):
     if request.method=="POST":
        
         form=studentform(request.POST)
         if form.is_valid():
             form.save()
         return render(request,'DOS/add.html',{
             'form': studentform(),
             'success': True
             })
     else:
         form=studentform()
     return render(request,'DOS/add.html',{
         'form': studentform()
                
             })
def edit(request,id):
  if request.method=='POST':
    student=student.objects.get(pk=id)
    form=studentform(request.POST,instance=student)
    if form.is_valid():
      form.save()
      return render(request,'DOS/edit.html',{
        'form':form,
        'success':True
      })
  else:
     student=student.objects.get(pk=id)
     form=studentform(instance=student)
  return render(request,'DOS/edit.html',{
     'form':form
  })
def delete(request,id):
  if request.method=='POST':
    student=student.objects.get(pk=id)
    student.delete()
  return HttpResponseRedirect(reverse('index'))

def theme(request):
    return render(request,'DOS/theme.html',{
        'DOS':student.objects.all()
    })
    
def pdf(request):
  buf= io.BytesIO()
  c = canvas.Canvas(buf,pagesize=letter,bottomup=0)
  textob = c.beginText()
  textob.setTextOrigin(inch,inch)
  textob.setFont("Helvetica",14)
  
  reportcards=reportcard.objects.all()

  lines =[]
  global line
  for reportcard in reportcards:
    lines.append(reportcard. std_No)
    lines.append(reportcard.subject)
    lines.append(reportcard. BOT)
    lines.append(reportcard. MOT)
    lines.append(reportcard. EOT)
    lines.append(reportcard.score)
    lines.append(reportcard.grade)
    lines.append(reportcard. comment)
    lines.append(reportcard. initials)
    lines.append(reportcard.totalmarks)
    lines.append(reportcard.averagemarks)
    lines.append("")
  
  for line in lines:
        textob.textLine(line)
  c.drawText(textob)
  c.showPage()
  c.save()
  buf.seek(0)  
  
  return FileResponse(buf,as_attachment=True,filename='pdf')

def generate_report_card(request):
    # Create a workbook and select active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    lines =[]
    global line
    
    reportcards=reportcard.objects.all()
    
    for reportcard in reportcards:
      lines.append(reportcard. std_No)
      lines.append(reportcard.subject)
      lines.append(reportcard. BOT)
      lines.append(reportcard. MOT)
      lines.append(reportcard. EOT)
      lines.append(reportcard.score)
      lines.append(reportcard.grade)
      lines.append(reportcard. comment)
      lines.append(reportcard. initials)
      lines.append(reportcard.totalmarks)
      lines.append(reportcard.averagemarks)
      lines.append("")
     
   
    for line in lines:
        sheet.append(line)
    
        col_letter = openpyxl.utils.get_column_letter(line,1)
        sheet['{}1'.format(col_letter)] = line
        sheet.column_dimensions[col_letter].width = 15  # adjust column width
        
    for row_num, student in enumerate(reportcards, 2):
        for col_num, field in enumerate(lines, 1):
            sheet.cell(row=row_num, column=col_num, value=student.get(field, ''))
        sheet.cell(row=row_num, column=len(line), value='=AVERAGE(B{}:D{})'.format(row_num, row_num))
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ReportCard.xlsx'
    
    workbook.save(response)

    return FileResponse(workbook,as_attachment=True,filename='excel')


